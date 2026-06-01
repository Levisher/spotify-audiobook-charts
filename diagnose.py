#!/usr/bin/env python3
"""Cross-verify: for each PDF, list every 'Audiobook · Author' OCR anchor and
the line directly above it (the title candidate), then mark which ones made it
into the xlsx."""
import re
import sys
from pathlib import Path

import pdfplumber
import pytesseract
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from spotify_scrape import ocr_page_lines, AUDIOBOOK_RE, clean_title, clean_author

XLSX = Path(__file__).parent / "audiobooks.xlsx"
DOWNLOADS = Path("/Users/keshavmittal/Downloads")
PLAYLIST_RE = re.compile(r"playlist-([A-Za-z0-9]{15,})")
SHEETID_RE = re.compile(r"^[A-Za-z0-9]{15,}$")


def pdf_for_sheet(sheet_id: str) -> Path | None:
    for p in DOWNLOADS.glob("screencapture-open-spotify-playlist-*.pdf"):
        m = PLAYLIST_RE.search(p.name)
        if m and m.group(1) == sheet_id:
            return p
    return None


def ocr_all_anchors(pdf_path: Path) -> list[dict]:
    """Return one dict per Audiobook anchor: {author, title_above, x, page}."""
    out = []
    with pdfplumber.open(pdf_path) as pdf:
        for pg_idx, page in enumerate(pdf.pages):
            im = page.to_image(resolution=200).original
            lines = ocr_page_lines(im)
            for i, ln in enumerate(lines):
                m = AUDIOBOOK_RE.match(ln["text"])
                if not m:
                    continue
                author = clean_author(m.group(1))
                # collect previous lines at same X within 120 px
                prevs = []
                for j in range(i - 1, -1, -1):
                    p = lines[j]
                    if p["y"] >= ln["y"]:
                        continue
                    if ln["y"] - p["y"] > 180:
                        break
                    prevs.insert(0, p)
                out.append({
                    "page": pg_idx,
                    "x": ln["x"],
                    "y": ln["y"],
                    "author": author,
                    "audiobook_line": ln["text"],
                    "prev_lines": [(p["x"], p["text"]) for p in prevs[-5:]],
                })
    return out


def xlsx_titles_for(sheet_id: str) -> list[tuple[str, str]]:
    wb = load_workbook(XLSX, read_only=True)
    if sheet_id not in wb.sheetnames:
        return []
    ws = wb[sheet_id]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for r in rows[1:]:
        out.append((str(r[0] or ""), str(r[1] or "")))
    return out


def main():
    wb = load_workbook(XLSX, read_only=True)
    sheets = [s for s in wb.sheetnames if SHEETID_RE.match(s)]
    for sn in sheets:
        pdf = pdf_for_sheet(sn)
        if not pdf:
            print(f"\n### {sn}  (PDF not found)\n")
            continue
        anchors = ocr_all_anchors(pdf)
        xlsx_rows = xlsx_titles_for(sn)
        xlsx_keys = {(t.lower(), a.lower()) for t, a in xlsx_rows}
        print(f"\n### {sn}  (OCR anchors: {len(anchors)}, xlsx rows: {len(xlsx_rows)})")
        print(f"   pdf: {pdf.name}")
        for k, a in enumerate(anchors):
            # what's "the title" using our parser's logic?
            parsed_title = ""
            for x, t in reversed(a["prev_lines"]):
                if abs(x - a["x"]) <= 40 and not AUDIOBOOK_RE.match(t):
                    parsed_title = clean_title(t)
                    break
            in_xlsx = any(a["author"].lower() == row_a.lower() for _, row_a in xlsx_rows)
            mark = "✓" if in_xlsx else "✗ MISSING"
            print(f"  [{k+1:2}] anchor=(x={a['x']:4}, y={a['y']:5})  page={a['page']}  author={a['author']!r}")
            print(f"       audiobook line: {a['audiobook_line']!r}")
            print(f"       prev lines:")
            for x, t in a["prev_lines"]:
                marker = " <-- same x" if abs(x - a["x"]) <= 40 else ""
                print(f"          x={x:4}  {t!r}{marker}")
            print(f"       parsed title  : {parsed_title!r}  {mark}")


if __name__ == "__main__":
    main()
