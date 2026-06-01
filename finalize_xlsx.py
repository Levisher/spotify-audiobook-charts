#!/usr/bin/env python3
"""
Final transforms on audiobooks.xlsx:
  1. Convert every date in 'First Published (print)' to DD/MM/YYYY.
  2. Rename 'Publisher (print)' -> 'Imprint (print)' on every sheet.
  3. Add 'Publishing House' column — only Big-5 + Bloomsbury named, rest 'Independent'.
  4. Add 'Backlist' column adjacent to the date — 1 if >12 months old (vs 2026-05-19).
  5. Add 'N' column at the end — total entries in that tab.
  6. Add a 'Summary' sheet with each tab's N, backlist count, backlist %.
"""
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

XLSX = Path(__file__).parent / "audiobooks.xlsx"
TODAY = date(2026, 5, 19)
CUTOFF = date(2025, 5, 19)  # >12 months old

# ---------- Imprint → Big-5 (or Bloomsbury) ----------
# Each (pattern, house). First match wins. Order matters — put more specific
# imprints before umbrella words that could overlap (e.g. "Penguin" before
# "Penguin Random House" is fine but "Harper" should come last among HC).
HOUSE_RULES: list[tuple[re.Pattern, str]] = [
    # Penguin Random House (PRH)
    (re.compile(r"penguin random house|\bprh\b", re.I), "Penguin Random House"),
    (re.compile(r"random house", re.I), "Penguin Random House"),
    (re.compile(r"\b(penguin|puffin|knopf|bantam|doubleday|vintage|anchor|"
                r"pantheon|riverhead|putnam|dutton|viking|tarcher|plume|"
                r"berkley|del rey|crown|hogarth|ballantine|fawcett|currency|"
                r"harmony|three rivers|dial press|schocken|spectra|avery|"
                r"perigee|dorling kindersley|spiegel\s*&\s*grau|one world|"
                r"modern library|pamela dorman|portfolio|sentinel)\b", re.I),
     "Penguin Random House"),

    # HarperCollins (HC)
    (re.compile(r"harpercollins|harper\s*collins", re.I), "HarperCollins"),
    (re.compile(r"\b(harperone|harper wave|harper voyager|harper perennial|"
                r"harpersanfrancisco|william morrow|mariner|avon|ecco|amistad|"
                r"witness|custom house|park row|\bmira\b|zondervan|"
                r"thomas nelson|\bhq\b|harper)\b", re.I),
     "HarperCollins"),

    # Simon & Schuster (S&S)
    (re.compile(r"simon\s*(?:&|and)\s*schuster|\bs\s*&\s*s\b", re.I),
     "Simon & Schuster"),
    (re.compile(r"\b(scribner|atria|gallery books?|touchstone|free press|"
                r"howard books|threshold editions|avid reader|saga press|"
                r"marysue rucci|tiller press|salaam reads)\b", re.I),
     "Simon & Schuster"),

    # Macmillan (MacM)
    (re.compile(r"macmillan", re.I), "Macmillan"),
    (re.compile(r"st\.?\s*martin'?s?", re.I), "Macmillan"),
    (re.compile(r"\b(henry holt|farrar\s+straus|\bfsg\b|flatiron|picador|"
                r"\btor\b|forge books?|bedford|holt paperbacks?|"
                r"first second|roaring brook|feiwel|priddy)\b", re.I),
     "Macmillan"),

    # Hachette (Hach)
    (re.compile(r"hachette", re.I), "Hachette"),
    (re.compile(r"little\s*,?\s*brown", re.I), "Hachette"),
    (re.compile(r"\b(grand central|orbit|mulholland|center street|faithwords|"
                r"hyperion|workman|algonquin|black dog\s*&\s*leventhal|"
                r"running press|bookouture|headline|\borion\b|mobius|"
                r"hodder\s*(?:&|and)?\s*stoughton|hodder|john murray|"
                r"perseus|public affairs|avalon|disney\s*hyperion)\b", re.I),
     "Hachette"),

    # Bloomsbury
    (re.compile(r"bloomsbury", re.I), "Bloomsbury"),
    (re.compile(r"\ba\s*&\s*c\s*black\b", re.I), "Bloomsbury"),

    # Historical PRH-acquired imprints
    (re.compile(r"george allen\s*&?\s*unwin|allen\s*&?\s*unwin", re.I),
     "HarperCollins"),  # Unwin Hyman -> HarperCollins 1990
]


def map_house(imprint: str) -> str:
    if not imprint:
        return "Independent"
    for pat, house in HOUSE_RULES:
        if pat.search(imprint):
            return house
    return "Independent"


# ---------- Date parsing ----------
def parse_date_cell(v) -> tuple[str, date | None]:
    """Return (DD/MM/YYYY display string, sortable date) or (original, None)."""
    if v is None:
        return ("", None)
    if isinstance(v, (datetime, date)):
        d = v.date() if isinstance(v, datetime) else v
        return (f"{d.day:02d}/{d.month:02d}/{d.year}", d)
    s = str(v).strip()
    if not s:
        return ("", None)
    # Already DD/MM/YYYY
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        d_, mo, y = map(int, m.groups())
        try:
            dd = date(y, mo, d_)
            return (f"{d_:02d}/{mo:02d}/{y}", dd)
        except ValueError:
            return (s, None)
    # ISO YYYY-MM-DD (possibly with time suffix)
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        y, mo, d_ = map(int, m.groups())
        try:
            dd = date(y, mo, d_)
            return (f"{d_:02d}/{mo:02d}/{y}", dd)
        except ValueError:
            return (s, None)
    # ISO YYYY-MM
    m = re.match(r"^(\d{4})-(\d{1,2})$", s)
    if m:
        y, mo = map(int, m.groups())
        try:
            dd = date(y, mo, 1)
            return (f"01/{mo:02d}/{y}", dd)
        except ValueError:
            return (s, None)
    # Bare year
    m = re.match(r"^(\d{4})$", s)
    if m:
        y = int(m.group(1))
        return (f"01/01/{y}", date(y, 1, 1))
    # Ancient text ("c. 5th century BCE")
    if "BCE" in s or re.search(r"\bcentury\b", s, re.I) or "BC" in s:
        return (s, date(1, 1, 1))
    return (s, None)


def is_backlist(d: date | None) -> bool:
    return d is not None and d < CUTOFF


# ---------- Transform a sheet ----------
HEADERS_TARGET = [
    "Title", "Author", "Imprint (print)", "Publishing House",
    "First Published (print)", "Backlist",
    "Matched Title", "Matched Authors", "Source", "N",
]


def transform_sheet(ws) -> tuple[int, int]:
    """Returns (n_rows, n_backlist)."""
    # Read existing header & rows
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return (0, 0)
    header = list(rows[0])

    # Identify columns by name in the original header
    col = {h: i for i, h in enumerate(header) if h is not None}

    def get(row, name):
        if name == "Imprint (print)" and name not in col:
            name = "Publisher (print)"
        i = col.get(name)
        return row[i] if i is not None and i < len(row) else None

    # Build new rows
    data_rows = []
    n_back = 0
    n = 0
    for r in rows[1:]:
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        title = get(r, "Title") or ""
        author = get(r, "Author") or ""
        imprint = get(r, "Imprint (print)") or get(r, "Publisher (print)") or ""
        date_raw = get(r, "First Published (print)")
        mt = get(r, "Matched Title") or ""
        ma = get(r, "Matched Authors") or ""
        src = get(r, "Source") or ""

        date_disp, parsed = parse_date_cell(date_raw)
        bl = 1 if is_backlist(parsed) else ""
        if bl == 1:
            n_back += 1
        n += 1

        house = map_house(str(imprint))
        data_rows.append([
            title, author, imprint, house,
            date_disp, bl,
            mt, ma, src, None,  # N filled after we know total
        ])

    # Fill N for every row
    for dr in data_rows:
        dr[-1] = n

    # Clear sheet & rewrite
    ws.delete_rows(1, ws.max_row)
    ws.append(HEADERS_TARGET)
    for dr in data_rows:
        ws.append(dr)

    # Format columns
    widths = {1: 42, 2: 24, 3: 26, 4: 22, 5: 14, 6: 10,
              7: 30, 8: 24, 9: 22, 10: 6}
    for cidx, w in widths.items():
        ws.column_dimensions[get_column_letter(cidx)].width = w
    ws.freeze_panes = "A2"
    return (n, n_back)


def write_summary(wb, stats: dict[str, tuple[int, int]]):
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary", 0)
    ws.append(["Tab", "N", "Backlist count", "Backlist %"])
    overall_n = 0
    overall_b = 0
    for name, (n, b) in stats.items():
        pct = (b / n * 100) if n else 0
        ws.append([name, n, b, round(pct, 1)])
        overall_n += n
        overall_b += b
    ws.append([])
    overall_pct = (overall_b / overall_n * 100) if overall_n else 0
    ws.append(["TOTAL", overall_n, overall_b, round(overall_pct, 1)])
    for cidx, w in {1: 28, 2: 8, 3: 16, 4: 12}.items():
        ws.column_dimensions[get_column_letter(cidx)].width = w
    ws.freeze_panes = "A2"


def main():
    wb = load_workbook(XLSX)
    stats: dict[str, tuple[int, int]] = {}
    # Snapshot sheet list BEFORE we touch anything; skip Summary if present
    sheets = [s for s in wb.sheetnames if s != "Summary"]
    for sn in sheets:
        ws = wb[sn]
        n, b = transform_sheet(ws)
        stats[sn] = (n, b)
        pct = (b / n * 100) if n else 0
        print(f"  {sn:30}  N={n:3}  backlist={b:3}  ({pct:.1f}%)")

    write_summary(wb, stats)
    wb.save(XLSX)
    print(f"\nSaved {XLSX.name}")


if __name__ == "__main__":
    main()
