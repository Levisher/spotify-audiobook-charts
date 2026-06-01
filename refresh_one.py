#!/usr/bin/env python3
"""
Refresh one playlist tab:
  1. Firecrawl https://open.spotify.com/playlist/{id}
  2. Parse N / Title / Author / banner from the rendered markdown
  3. Gemini lookup for each book -> first-edition publisher + date
  4. Map imprint -> Big-5 / Bloomsbury / Independent
  5. Compute Backlist (>12 mo old vs today)
  6. Update audiobooks.xlsx tab in place (or create it) — keeps existing rows
     that aren't in the new fetch by marking them inactive in Source-like col.
  7. Render a PDF of the new tab into ~/Downloads/

Usage:
    python3 refresh_one.py 37i9dQZF1CHSUrTnfwXo9Y
"""
import json
import os
import re
import subprocess
import sys
import time
from datetime import date, datetime
from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                TableStyle)

ROOT = Path(__file__).parent
CREDS = ROOT / ".creds"
XLSX = ROOT / "audiobooks.xlsx"
# PDFs go to $OUTPUT_DIR if set (used in CI), else ~/Downloads locally.
DOWNLOADS = Path(os.environ["OUTPUT_DIR"]) if os.environ.get("OUTPUT_DIR") \
    else Path.home() / "Downloads"
TODAY = date.today()
CUTOFF = date(TODAY.year - 1, TODAY.month, TODAY.day)

GEMINI_KEY = (CREDS / "gemini_api_key.txt").read_text().strip()
GEMINI_URL = (
    "https://generativelanguage.googleapis.com/v1beta/"
    "models/gemini-2.5-flash:generateContent?key=" + GEMINI_KEY
)


# ---------- Firecrawl ----------
def get_firecrawl_key():
    # CI sets FIRECRAWL_API_KEY directly. Local Mac uses the firecrawl-key
    # CLI swapper that auto-rotates across multiple keys.
    if os.environ.get("FIRECRAWL_API_KEY"):
        return os.environ["FIRECRAWL_API_KEY"].strip()
    return subprocess.check_output(["firecrawl-key"], text=True).strip()


def firecrawl_scrape(url: str) -> str:
    key = get_firecrawl_key()
    r = requests.post(
        "https://api.firecrawl.dev/v2/scrape",
        headers={"Authorization": f"Bearer {key}",
                 "Content-Type": "application/json"},
        json={"url": url, "formats": ["markdown"],
              "onlyMainContent": False, "waitFor": 3000},
        timeout=120,
    )
    r.raise_for_status()
    d = r.json()
    if not d.get("success"):
        raise RuntimeError(f"Firecrawl failed: {d}")
    return d["data"]["markdown"]


# ---------- Parse Spotify playlist markdown ----------
AUDIOBOOK_LINE = re.compile(r"^Audiobook\s*[•·]\s*(.+?)\s*$", re.IGNORECASE)


def parse_playlist(md: str) -> tuple[str, list[dict]]:
    """Returns (banner_name, [{N,title,author}, ...]).
    The markdown layout is:
        # <Banner>
        Top <X> audiobooks, updated weekly
        ...
        N         <- standalone number line
        <Title>
        Audiobook • <Author>
        <description...>
    """
    # Banner: prefer the # heading that appears after the navigation block.
    # In practice the H1 we want is the playlist name. Take the LAST H1
    # before the first 'Audiobook •' line.
    lines = [ln.rstrip() for ln in md.splitlines()]
    first_book_i = None
    for i, ln in enumerate(lines):
        if AUDIOBOOK_LINE.match(ln):
            first_book_i = i
            break
    banner = ""
    if first_book_i:
        for j in range(first_book_i - 1, -1, -1):
            ln = lines[j].strip()
            m = re.match(r"^#\s+(.+)$", ln)
            if m and "your library" not in m.group(1).lower():
                banner = m.group(1).strip()
                break

    def unescape_md(s: str) -> str:
        # Firecrawl emits markdown escapes for special chars: \* \_ \# etc.
        return re.sub(r"\\([\*_#\[\]\(\)`~\\])", r"\1", s)

    entries: list[dict] = []
    for i, ln in enumerate(lines):
        m = AUDIOBOOK_LINE.match(ln.strip())
        if not m:
            continue
        author = unescape_md(m.group(1).strip())
        # Walk back for title (nearest non-empty line that isn't itself
        # an Audiobook line or a pure number).
        title = ""
        serial = None
        for j in range(i - 1, max(i - 8, -1), -1):
            cand = lines[j].strip()
            if not cand:
                continue
            if AUDIOBOOK_LINE.match(cand):
                continue
            if re.fullmatch(r"\d+", cand):
                serial = int(cand)
                break
            if not title:
                title = cand
        if title:
            entries.append({
                "n": serial if serial is not None else len(entries) + 1,
                "title": unescape_md(title),
                "author": author,
            })
    return banner, entries


# ---------- Gemini lookup (with disk cache) ----------
GEM_CACHE = ROOT / ".gemini_cache.json"
_gem_cache = json.loads(GEM_CACHE.read_text()) if GEM_CACHE.exists() else {}


def _save_gem_cache():
    GEM_CACHE.write_text(json.dumps(_gem_cache, indent=2))


def gemini_lookup(title: str, author: str) -> dict:
    key = f"{title}::{author}"
    if key in _gem_cache:
        return _gem_cache[key]
    prompt = (
        f'Book: "{title}" by {author}.\n'
        'Return ONLY one JSON object: '
        '{"publisher": "<first-edition print publisher>", '
        '"date": "<YYYY-MM-DD or YYYY-MM or YYYY>", '
        '"confidence": "high|medium|low"}\n'
        'Use "N/A" for publisher or date when unknown. '
        'No prose, no markdown, no code fences.'
    )
    body = {
        "contents": [{"parts": [{"text": prompt}]}],
        "tools": [{"google_search": {}}],
        "generationConfig": {
            "temperature": 0.0,
            "maxOutputTokens": 800,
            "thinkingConfig": {"thinkingBudget": 0},
        },
    }
    r = requests.post(GEMINI_URL, json=body, timeout=45)
    if r.status_code != 200:
        return {"publisher": "", "date": "", "confidence": "error",
                "_error": f"HTTP {r.status_code}: {r.text[:200]}"}
    cand = r.json().get("candidates", [{}])[0]
    parts = cand.get("content", {}).get("parts", [])
    text = "".join(p.get("text", "") for p in parts).strip()
    # strip possible code fences
    text = re.sub(r"^```(?:json)?|```$", "", text, flags=re.MULTILINE).strip()
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        m = re.search(r"\{.+\}", text, re.DOTALL)
        if not m:
            return {"publisher": "", "date": "", "confidence": "parse_error",
                    "_raw": text[:200]}
        try:
            data = json.loads(m.group(0))
        except Exception:
            return {"publisher": "", "date": "", "confidence": "parse_error",
                    "_raw": text[:200]}
    pub = data.get("publisher", "") or ""
    dt = data.get("date", "") or ""
    if pub.upper() == "N/A":
        pub = ""
    if dt.upper() == "N/A":
        dt = ""
    result = {"publisher": pub, "date": dt,
              "confidence": data.get("confidence", "")}
    _gem_cache[key] = result
    _save_gem_cache()
    return result


# ---------- House mapping (same rules as finalize_xlsx.py) ----------
HOUSE_RULES = [
    (re.compile(r"penguin random house|\bprh\b", re.I), "Penguin Random House"),
    (re.compile(r"random house", re.I), "Penguin Random House"),
    (re.compile(
        r"\b(penguin|puffin|knopf|bantam|doubleday|vintage|anchor|pantheon|"
        r"riverhead|putnam|dutton|viking|tarcher|plume|berkley|del rey|crown|"
        r"hogarth|ballantine|fawcett|currency|harmony|three rivers|dial press|"
        r"schocken|spectra|avery|perigee|dorling kindersley|portfolio|sentinel)\b",
        re.I), "Penguin Random House"),
    (re.compile(r"harpercollins|harper\s*collins", re.I), "HarperCollins"),
    (re.compile(
        r"\b(harperone|harper wave|harper voyager|harper perennial|"
        r"harpersanfrancisco|william morrow|mariner|avon|ecco|amistad|witness|"
        r"custom house|park row|mira|zondervan|thomas nelson|hq|harper)\b",
        re.I), "HarperCollins"),
    (re.compile(r"simon\s*(?:&|and)\s*schuster|\bs\s*&\s*s\b", re.I),
     "Simon & Schuster"),
    (re.compile(
        r"\b(scribner|atria|gallery books?|touchstone|free press|howard books|"
        r"threshold editions|avid reader|saga press|marysue rucci|tiller press)\b",
        re.I), "Simon & Schuster"),
    (re.compile(r"macmillan", re.I), "Macmillan"),
    (re.compile(r"st\.?\s*martin'?s?", re.I), "Macmillan"),
    (re.compile(
        r"\b(henry holt|farrar\s+straus|\bfsg\b|flatiron|picador|\btor\b|"
        r"forge books?|bedford|holt paperbacks?|first second|roaring brook|feiwel)\b",
        re.I), "Macmillan"),
    (re.compile(r"hachette", re.I), "Hachette"),
    (re.compile(r"little\s*,?\s*brown", re.I), "Hachette"),
    (re.compile(
        r"\b(grand central|orbit|mulholland|center street|faithwords|hyperion|"
        r"workman|algonquin|black dog\s*&\s*leventhal|running press|bookouture|"
        r"headline|orion|mobius|hodder|john murray|perseus|public affairs|"
        r"disney hyperion)\b", re.I), "Hachette"),
    (re.compile(r"bloomsbury", re.I), "Bloomsbury"),
    (re.compile(r"\ba\s*&\s*c\s*black\b", re.I), "Bloomsbury"),
    (re.compile(r"george allen\s*&?\s*unwin|allen\s*&?\s*unwin", re.I),
     "HarperCollins"),
]


def map_house(imprint: str) -> str:
    if not imprint:
        return "Independent"
    for pat, h in HOUSE_RULES:
        if pat.search(imprint):
            return h
    return "Independent"


# ---------- Date helpers ----------
def parse_iso(s: str):
    s = (s or "").strip()
    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def ddmmyyyy(d):
    if not d:
        return ""
    return f"{d.day:02d}/{d.month:02d}/{d.year}"


def is_backlist(d):
    return d is not None and d < CUTOFF


# ---------- xlsx ops ----------
HEADERS = ["N", "Title", "Author", "Imprint (print)", "Publishing House",
           "First Published (print)", "Backlist"]


def _author_match(a1: str, a2: str) -> bool:
    """Loose match: last name token shared."""
    def last(a):
        toks = re.findall(r"[A-Za-z]+", a)
        return toks[-1].lower() if toks else ""
    return last(a1) == last(a2) and last(a1) != ""


def _title_match(t1: str, t2: str) -> bool:
    """Match on the title head (pre-colon, pre-em-dash)."""
    def head(t):
        return re.split(r"[:—\-]", t, maxsplit=1)[0].strip().lower()
    return head(t1) == head(t2) and head(t1) != ""


def diff_against_existing(banner: str, new_rows: list[dict],
                          source_xlsx: Path | None = None) -> dict:
    """Returns rich diff with previous ranks. Match on (title-head, last-name)
    so OCR'd Spotify subtitle variants still align."""
    path = source_xlsx or XLSX
    wb = load_workbook(path) if path.exists() else None
    prev_entries: list[dict] = []
    if wb and banner in wb.sheetnames:
        ws = wb[banner]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = list(rows[0])
            col = {h: i for i, h in enumerate(header) if h}
            for r in rows[1:]:
                if not r:
                    continue
                t = str(r[col.get("Title", 1)] or "").strip()
                a = str(r[col.get("Author", 2)] or "").strip()
                n = r[col.get("N", 0)]
                if t and a:
                    prev_entries.append({"n": n, "title": t, "author": a})

    # For each new row, find a matching previous entry
    matched_prev_idx = set()
    for i, nr in enumerate(new_rows):
        nr["prev_n"] = None
        for j, pe in enumerate(prev_entries):
            if j in matched_prev_idx:
                continue
            if (_author_match(nr["author"], pe["author"])
                    and _title_match(nr["title"], pe["title"])):
                nr["prev_n"] = pe["n"]
                matched_prev_idx.add(j)
                break
    dropped = [pe for j, pe in enumerate(prev_entries) if j not in matched_prev_idx]
    return {
        "added": [r for r in new_rows if r["prev_n"] is None],
        "dropped": dropped,
        "unchanged": [r for r in new_rows if r["prev_n"] is not None],
    }


def write_tab(banner: str, rows: list[dict]):
    wb = load_workbook(XLSX) if XLSX.exists() else None
    if wb is None:
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)
    if banner in wb.sheetnames:
        del wb[banner]
    ws = wb.create_sheet(banner)
    ws.append(HEADERS)
    for i, r in enumerate(rows, start=1):
        ws.append([
            i, r["title"], r["author"],
            r["imprint"], r["house"],
            ddmmyyyy(r["date_obj"]),
            1 if is_backlist(r["date_obj"]) else "",
        ])
    for cidx, w in enumerate([6, 42, 24, 26, 22, 16, 10], start=1):
        ws.column_dimensions[get_column_letter(cidx)].width = w
    ws.freeze_panes = "A2"
    wb.save(XLSX)


# ---------- PDF rendering ----------
def rank_delta(prev_n, cur_n) -> tuple[str, str]:
    """Returns (display, color_hex) for the Δ column."""
    if prev_n is None:
        return ("NEW", "#1a8917")  # green
    try:
        d = int(prev_n) - int(cur_n)
    except (TypeError, ValueError):
        return ("—", "#666666")
    if d == 0:
        return ("—", "#666666")
    if d > 0:
        return (f"↑{d}", "#1a8917")  # rose up
    return (f"↓{-d}", "#c62828")     # fell down


def render_pdf(banner: str, rows: list[dict], dropped: list[dict],
               out_path: Path):
    doc = SimpleDocTemplate(
        str(out_path), pagesize=landscape(letter),
        leftMargin=0.4 * inch, rightMargin=0.4 * inch,
        topMargin=0.5 * inch, bottomMargin=0.5 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"], fontSize=18, alignment=0,
        spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        "Sub", parent=styles["Normal"], fontSize=9, textColor=colors.grey,
        spaceAfter=12,
    )
    section = ParagraphStyle(
        "Section", parent=styles["Heading3"], fontSize=11,
        spaceBefore=14, spaceAfter=4,
    )

    n_total = len(rows)
    n_backlist = sum(1 for r in rows if is_backlist(r["date_obj"]))
    n_new = sum(1 for r in rows if r.get("prev_n") is None)
    pct = (n_backlist / n_total * 100) if n_total else 0

    story = [
        Paragraph(banner, title_style),
        Paragraph(
            f"{n_total} books • Backlist: {n_backlist} ({pct:.1f}%) • "
            f"New entries: {n_new} • Dropped: {len(dropped)} • "
            f"Refreshed {TODAY.strftime('%d %b %Y')}",
            sub_style,
        ),
    ]

    cell = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=9,
                          leading=11)
    data = [["N", "Δ", "Prev", "Title", "Author", "Imprint",
             "Publishing House", "First Published", "BL"]]

    delta_styles: list[tuple] = []
    for i, r in enumerate(rows, start=1):
        delta_text, delta_color = rank_delta(r.get("prev_n"), i)
        delta_styles.append(
            ("TEXTCOLOR", (1, i), (1, i), colors.HexColor(delta_color))
        )
        delta_styles.append(
            ("FONTNAME", (1, i), (1, i), "Helvetica-Bold")
        )
        if delta_text == "NEW":
            delta_styles.append(
                ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#e7f6e7"))
            )
        data.append([
            str(i),
            delta_text,
            str(r.get("prev_n") or "—"),
            Paragraph(r["title"], cell),
            Paragraph(r["author"], cell),
            Paragraph(str(r["imprint"] or ""), cell),
            r["house"],
            ddmmyyyy(r["date_obj"]),
            "1" if is_backlist(r["date_obj"]) else "",
        ])
    col_widths = [0.35 * inch, 0.45 * inch, 0.45 * inch, 2.6 * inch,
                  1.4 * inch, 1.7 * inch, 1.25 * inch, 1.0 * inch, 0.35 * inch]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    base_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#222")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 1), (2, -1), "CENTER"),
        ("ALIGN", (-2, 1), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#f7f7f7")]),
    ]
    t.setStyle(TableStyle(base_style + delta_styles))
    story.append(t)

    if dropped:
        story.append(Paragraph(
            f"Dropped since last refresh ({len(dropped)})", section))
        dropped_data = [["Prev N", "Title", "Author"]]
        for d in dropped:
            dropped_data.append([
                str(d.get("n") or ""),
                Paragraph(str(d.get("title") or ""), cell),
                Paragraph(str(d.get("author") or ""), cell),
            ])
        dt = Table(dropped_data,
                   colWidths=[0.6 * inch, 4.5 * inch, 2.0 * inch],
                   repeatRows=1)
        dt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7a0e0e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 1), (0, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cccccc")),
        ]))
        story.append(dt)

    doc.build(story)


# ---------- Main ----------
def main(playlist_id: str, use_cache: bool = False,
         tab_override: str | None = None):
    cache = ROOT / f".cache_{playlist_id}.md"
    if use_cache and cache.exists():
        print(f"[cache] reading {cache.name}", file=sys.stderr)
        md = cache.read_text()
    else:
        url = f"https://open.spotify.com/playlist/{playlist_id}"
        print(f"[firecrawl] {url}", file=sys.stderr)
        md = firecrawl_scrape(url)
        cache.write_text(md)
        print(f"  cached -> {cache.name} ({len(md)} chars)", file=sys.stderr)

    banner, entries = parse_playlist(md)
    if tab_override:
        if banner and banner != tab_override:
            print(f"  banner from page: {banner!r} (overridden -> {tab_override!r})",
                  file=sys.stderr)
        banner = tab_override
    if not banner:
        banner = f"Playlist {playlist_id[:10]}"
    print(f"  banner: {banner!r}, entries: {len(entries)}", file=sys.stderr)

    rows = []
    for e in entries:
        print(f"  [gemini] {e['title'][:50]!r} / {e['author']!r}",
              file=sys.stderr)
        meta = gemini_lookup(e["title"], e["author"])
        date_obj = parse_iso(meta.get("date", ""))
        imprint = meta.get("publisher", "")
        rows.append({
            "title": e["title"],
            "author": e["author"],
            "imprint": imprint,
            "house": map_house(imprint),
            "date_obj": date_obj,
            "date_raw": meta.get("date", ""),
            "confidence": meta.get("confidence", ""),
        })
        time.sleep(0.4)

    # Diff vs the pre-refresh snapshot (so re-runs don't show a no-op diff)
    snap = ROOT / "audiobooks.before_refresh.xlsx"
    src = snap if snap.exists() else XLSX
    diff = diff_against_existing(banner, rows, source_xlsx=src)
    print(f"\n[diff] added: {len(diff['added'])}, "
          f"dropped: {len(diff['dropped'])}, "
          f"unchanged: {len(diff['unchanged'])}", file=sys.stderr)
    for r in rows:
        if r["prev_n"] is None:
            print(f"  + NEW   {r['title']!r:40.40} / {r['author']!r}",
                  file=sys.stderr)
    for d in diff["dropped"]:
        print(f"  -       {d['title']!r:40.40} / {d['author']!r}",
              file=sys.stderr)
    for r in rows:
        if r["prev_n"] is not None:
            try:
                delta = int(r["prev_n"]) - rows.index(r) - 1
                if delta != 0:
                    arrow = "↑" if delta > 0 else "↓"
                    print(f"  {arrow}{abs(delta):<5} {r['title']!r:40.40}",
                          file=sys.stderr)
            except (TypeError, ValueError):
                pass

    # Write xlsx + PDF
    write_tab(banner, rows)
    pdf_name = (
        f"audiobooks_{banner.replace(' ', '_').replace('&','and')}_"
        f"{TODAY.isoformat()}.pdf"
    )
    pdf_path = DOWNLOADS / pdf_name
    render_pdf(banner, rows, diff["dropped"], pdf_path)
    print(f"\n[done] xlsx tab: {banner!r}", file=sys.stderr)
    print(f"[done] pdf: {pdf_path}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("playlist_id")
    p.add_argument("--cache", action="store_true")
    p.add_argument("--tab", default=None,
                   help="Override the destination tab name (e.g. preserve a manual rename)")
    a = p.parse_args()
    raise SystemExit(main(a.playlist_id, use_cache=a.cache, tab_override=a.tab))
