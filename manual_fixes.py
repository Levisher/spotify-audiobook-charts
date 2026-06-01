#!/usr/bin/env python3
"""Targeted corrections after auto-verification:
  - Revert deep_verify regressions where an earlier-date wrong-book match landed
  - Apply known canonical first editions for famous books
  - Strip OCR cosmetic typos in titles
"""
import re
from pathlib import Path
from openpyxl import load_workbook

XLSX = Path(__file__).parent / "audiobooks.xlsx"

# (sheet, title-substring, [new_title, new_author, new_pub, new_date])
# None = keep existing
FIXES: list[tuple[str, str, dict]] = [
    # ---- Regressions from deep_verify ----
    ("History", "Third Reich",
     {"publisher": "Simon & Schuster", "date": "2017-10-10",
      "matched_title": "The Third Reich", "matched_authors": "Thomas Childers"}),
    ("Parenting & Relationships", "Men with Adult ADHD",
     {"date": "2023-01-28"}),
    ("Business & Careers", "Think and Grow Rich!",
     {"publisher": "Tarcher", "date": "2005-08-18",
      "matched_title": "Think and Grow Rich: The Original 1937 Classic",
      "matched_authors": "Napoleon Hill"}),
    ("Business & Careers", "Think and Grow Rich",  # the non-! one
     {"publisher": "The Ralston Society", "date": "1937",
      "matched_title": "Think and Grow Rich",
      "matched_authors": "Napoleon Hill"}),

    # ---- Canonical first editions ----
    ("All Fantasy", "Fellowship of the Ring",
     {"publisher": "George Allen & Unwin", "date": "1954-07-29"}),
    ("All Fantasy", "A Game of Thrones",
     {"publisher": "Bantam Spectra", "date": "1996-08-01"}),
    ("All Fantasy", "Dune: Book One",
     {"publisher": "Chilton Books", "date": "1965-08"}),
    ("All Fantasy", "A Court of Wings and Ruin",
     {"publisher": "Bloomsbury Publishing", "date": "2017-05-02"}),

    ("Teen & Young Adult", "Twilight",
     {"publisher": "Little, Brown and Company", "date": "2005-10-05"}),
    ("Teen & Young Adult", "Lord of the Flies",
     {"publisher": "Faber & Faber", "date": "1954-09-17"}),
    ("Teen & Young Adult", "Summer | Turned Pretty",
     {"title": "The Summer I Turned Pretty",
      "publisher": "Simon & Schuster Books for Young Readers", "date": "2009-05-05"}),
    ("Teen & Young Adult", "Eragon: Inheritance, Book",
     {"title": "Eragon: Inheritance, Book 1",
      "publisher": "Knopf Books for Young Readers", "date": "2003-08-26"}),

    ("Business & Careers", "7 Habits of Highly Effective",
     {"publisher": "Free Press", "date": "1989-08-15"}),

    ("Happiness & Success", "Atomic Habits",
     {"publisher": "Avery", "date": "2018-10-16"}),
    ("Happiness & Success", "Subtle Art of Not Giving",
     {"publisher": "HarperOne", "date": "2016-09-13"}),
    ("Happiness & Success", "How To Win Friends",
     {"publisher": "Simon & Schuster", "date": "1936-10-01"}),

    ("History", "Empire of the Summer Moon",
     {"publisher": "Scribner", "date": "2010-05-25"}),
    ("History", "Nation",  # Sorkin
     {"title": "1929: Inside the Greatest Crash in Wall Street History—and How It Shattered a Nation",
      "publisher": "Penguin Press", "date": "2025-09-30",
      "matched_title": "1929",
      "matched_authors": "Andrew Ross Sorkin"}),

    ("Entertainment Biography", "I'm Glad My Mom Died",
     {"publisher": "Simon & Schuster", "date": "2022-08-09"}),

    ("Mystery & Thriller", "Housemaid's Secret",
     {"publisher": "Grand Central Publishing", "date": "2022-10-25"}),
    ("Mystery & Thriller", "Housemaid Is Watching",
     {"publisher": "Grand Central Publishing", "date": "2024-06-25"}),

    ("Religion & Spirituality", "Becoming Nobody",
     {"publisher": "Sounds True", "date": "2019-04-30"}),
    ("Religion & Spirituality", "Four Agreements",
     {"publisher": "Amber-Allen Publishing", "date": "1997-11-07"}),

    ("Parenting & Relationships", "to Resilient Parenting",
     {"title": "Good Inside: A Guide to Becoming the Parent You Want to Be",
      "publisher": "Harper Wave", "date": "2022-09-13",
      "matched_title": "Good Inside",
      "matched_authors": "Becky Kennedy"}),

    # ---- Minor OCR title typos ----
    ("Entertainment Biography", "AMemoir",
     {"title": "You with the Sad Eyes: A Memoir"}),

    # ---- Precision upgrades for year-only dates ----
    ("History", "Mythos", {"date": "2017-11-02"}),
    ("All Fantasy", "Crown of Midnight",
     {"publisher": "Bloomsbury", "date": "2013-08-27"}),
    ("All Fantasy", "Throne of Glass",
     {"publisher": "Bloomsbury", "date": "2012-08-07"}),
    ("All Fantasy", "Heir of Fire",
     {"publisher": "Bloomsbury", "date": "2014-09-02"}),
    ("All Fantasy", "Fourth Wing",
     {"publisher": "Entangled: Red Tower Books", "date": "2023-05-02"}),
    ("All Fantasy", "Iron Flame",
     {"publisher": "Entangled: Red Tower Books", "date": "2023-11-07"}),
    ("History", "The Art of War",
     {"publisher": "Createspace Independent Publishing Platform", "date": "c. 5th century BCE"}),

    # K Pop -> KPop (Spotify's actual display)
    ("Kids & Family", "K Pop Demon Hunters",
     {"title": "KPop Demon Hunters: The Official Deluxe Junior Novelization"}),
    ("Kids & Family", "KPop Demon Hunters",
     {"title": "KPop Demon Hunters: The Official Deluxe Junior Novelization"}),

    # ---- Revert remaining deep_verify regressions ----
    ("Happiness & Success", "Power of Now",
     {"publisher": "New World Library", "date": "1999-09-27"}),
    ("Happiness & Success", "Stop Letting Everything Affect You",
     {"date": "2025-08-13"}),  # publisher remains as deep_verify set it
    ("History", "We the People",
     {"publisher": "Liveright", "date": "2025-09-16",
      "matched_authors": "Jill Lepore"}),  # Liveright = US first ed
    ("Religion & Spirituality", "Alchemist",
     {"publisher": "HarperOne", "date": "1993-04-25"}),  # US first ed
]


def apply_fixes():
    wb = load_workbook(XLSX)
    applied = 0
    for sheet_name, substr, changes in FIXES:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        col = {h: i for i, h in enumerate(header)}
        matched = False
        for row in ws.iter_rows(min_row=2):
            t = str(row[col["Title"]].value or "")
            if substr.lower() not in t.lower():
                continue
            if "title" in changes:
                row[col["Title"]].value = changes["title"]
            if "author" in changes:
                row[col["Author"]].value = changes["author"]
            if "publisher" in changes:
                row[col["Publisher (print)"]].value = changes["publisher"]
            if "date" in changes:
                row[col["First Published (print)"]].value = changes["date"]
            if "matched_title" in changes and "Matched Title" in col:
                row[col["Matched Title"]].value = changes["matched_title"]
            if "matched_authors" in changes and "Matched Authors" in col:
                row[col["Matched Authors"]].value = changes["matched_authors"]
            if "Source" in col:
                cur = str(row[col["Source"]].value or "")
                row[col["Source"]].value = cur + (";manual" if "manual" not in cur else "")
            print(f"[{sheet_name}] {t[:40]} -> {changes}")
            applied += 1
            matched = True
            break
        if not matched:
            print(f"  ! no row matched substr {substr!r} in {sheet_name}")

    wb.save(XLSX)
    print(f"\nApplied {applied} fixes")


if __name__ == "__main__":
    apply_fixes()
