#!/usr/bin/env python3
"""
For each row (non-preserved sheet), look at ALL Google Books editions that
match the author + title, filter out junk publishers (reprints, large-print,
self-pub, foreign reprints when an original US/UK pub exists), and pick the
earliest dated edition with a real publisher. If that edition's date is
strictly earlier than what's currently in the cell, update both publisher
and date.

This catches:
  - Atomic Habits / Manjul Publishing / 2018  ->  Avery / 2018-10-16
  - Dune / Penguin / 1975                     ->  Chilton / 1965-08
  - Fellowship / Del Rey / 1965               ->  Allen & Unwin / 1954-07
  - A Game of Thrones / HarperCollins UK / 2010 -> Bantam / 1996-08
"""
import re
import sys
import time
from pathlib import Path

import requests
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from spotify_scrape import GOOGLE_BOOKS, _get_with_retries, _author_matches, _short_title

PRESERVE = {"Top", "Romance"}

# Publishers that strongly imply a reprint / large-print / foreign / aggregator.
# When we see these alongside an original-publisher edition with an earlier
# date, prefer the original.
JUNK_PUBLISHERS = {
    "thorndike press large print", "thorndike press", "createspace",
    "createspace independent publishing platform", "scb distributors",
    "strelbytskyy multimedia publishing", "francisgalton",
    "national geographic books",  # ghosted on Junie B
    "paw prints",  # repackager
    "turtleback books",  # library binding rebinder
}


def is_junk(pub: str) -> bool:
    return (pub or "").strip().lower() in JUNK_PUBLISHERS


def normalize_words(t: str) -> set[str]:
    return {w.lower() for w in re.findall(r"\b[a-zA-Z]{4,}\b", t)}


def title_overlap(a: str, b: str) -> float:
    sa, sb = normalize_words(a), normalize_words(b)
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / max(len(sa), 1)


def find_best_edition(title: str, author: str, session, api_key: str) -> dict | None:
    short = _short_title(title)
    queries = [
        {"q": f'intitle:"{short}" inauthor:"{author}"'},
        {"q": f'"{short}" "{author}"'},
        {"q": f"{short} {author}"},
    ]
    all_items: list[dict] = []
    for q in queries:
        params = {**q, "printType": "books", "maxResults": 40, "key": api_key}
        try:
            data = _get_with_retries(session, GOOGLE_BOOKS, params)
        except Exception:
            continue
        items = data.get("items") or []
        all_items.extend(items)
        if len(all_items) >= 30:
            break
    if not all_items:
        return None

    a_low = author.lower()
    # Filter: author must match AND title must share enough words
    matches = []
    for it in all_items:
        vi = it.get("volumeInfo", {})
        if not _author_matches(a_low, vi.get("authors") or []):
            continue
        if title_overlap(short, vi.get("title", "")) < 0.4:
            continue
        if not vi.get("publishedDate"):
            continue
        matches.append(it)
    if not matches:
        return None

    # Prefer non-junk publishers with earliest date
    real = [it for it in matches
            if it.get("volumeInfo", {}).get("publisher")
            and not is_junk(it["volumeInfo"]["publisher"])]
    if not real:
        real = matches  # all junk — at least pick earliest

    real.sort(key=lambda it: it["volumeInfo"].get("publishedDate", "9999"))
    return real[0]


def main(xlsx_path: str):
    p = Path(xlsx_path)
    wb = load_workbook(p)

    api_key_file = Path(__file__).parent / "google_books_key.txt"
    api_key = api_key_file.read_text().strip() if api_key_file.exists() else ""
    session = requests.Session()

    upgrades = 0
    for sn in wb.sheetnames:
        if sn in PRESERVE:
            continue
        ws = wb[sn]
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        col = {h: i for i, h in enumerate(header)}
        i_title = col["Title"]
        i_auth = col["Author"]
        i_pub = col["Publisher (print)"]
        i_date = col["First Published (print)"]
        i_mt = col.get("Matched Title", -1)
        i_ma = col.get("Matched Authors", -1)
        i_src = col.get("Source", -1)

        for row in ws.iter_rows(min_row=2):
            title = str(row[i_title].value or "")
            author = str(row[i_auth].value or "")
            cur_pub = str(row[i_pub].value or "")
            cur_date = str(row[i_date].value or "")
            if not title or not author:
                continue

            best = find_best_edition(title, author, session, api_key)
            time.sleep(0.15)
            if not best:
                continue
            vi = best["volumeInfo"]
            new_pub = vi.get("publisher", "") or cur_pub
            new_date = vi.get("publishedDate", "") or cur_date

            # Only upgrade if strictly better:
            # (a) new_date is non-empty AND strictly earlier than cur_date, OR
            # (b) cur publisher is junk and new publisher is real
            cur_is_junk = is_junk(cur_pub)
            cur_d = cur_date or "9999"
            improved = False
            if new_date and new_date < cur_d and len(new_date) >= 4:
                improved = True
            elif cur_is_junk and new_pub and not is_junk(new_pub):
                improved = True
            elif not cur_pub and new_pub:  # fill in blank publisher
                improved = True

            if not improved:
                continue

            old = f"{cur_pub!r} / {cur_date!r}"
            new = f"{new_pub!r} / {new_date!r}"
            print(f"[{sn}] {title[:40]!r:42} | {old}  ->  {new}", file=sys.stderr)
            row[i_pub].value = new_pub
            row[i_date].value = new_date
            if i_mt >= 0:
                row[i_mt].value = vi.get("title", "")
            if i_ma >= 0:
                row[i_ma].value = ", ".join(vi.get("authors") or [])
            if i_src >= 0:
                cur_src = str(row[i_src].value or "")
                row[i_src].value = cur_src + (";deep_verify" if "deep_verify" not in cur_src else "")
            upgrades += 1

    wb.save(p)
    print(f"\nDone. Upgrades: {upgrades}", file=sys.stderr)


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "audiobooks.xlsx")
