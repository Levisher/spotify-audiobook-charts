#!/usr/bin/env python3
"""
Spotify audiobook playlist scraper.

Reads one or more PDF screencaptures of a Spotify audiobook playlist, OCRs each
page, extracts Title + Author pairs, then enriches each row with Publisher
and First Published date from the Google Books API. Writes an .xlsx where
each input PDF becomes its own worksheet tab.

Usage:
    python3 spotify_scrape.py OUT.xlsx PDF [PDF ...]
"""

import argparse
import re
import sys
import time
from pathlib import Path

import pandas as pd
import pdfplumber
import pytesseract
import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

GOOGLE_BOOKS = "https://www.googleapis.com/books/v1/volumes"
OPEN_LIBRARY = "https://openlibrary.org/search.json"
AUDIOBOOK_RE = re.compile(r"^\s*Audiobook\b[\s\W]+(.+?)\s*$", re.IGNORECASE)


def ocr_page_lines(image) -> list[dict]:
    """Return one dict per OCR line with keys: text, x, y, w, h.
    x = leftmost word's left; y = line top; h = max word height in the line.
    """
    df = pytesseract.image_to_data(image, output_type=pytesseract.Output.DATAFRAME)
    df = df.dropna(subset=["text"])
    df["text"] = df["text"].astype(str)
    df = df[df["text"].str.strip() != ""]
    lines = []
    grp_cols = ["page_num", "block_num", "par_num", "line_num"]
    for keys, sub in df.groupby(grp_cols, sort=False):
        sub = sub.sort_values("left")
        txt = " ".join(sub["text"].tolist()).strip()
        if not txt:
            continue
        lines.append({
            "text": txt,
            "x": int(sub["left"].min()),
            "y": int(sub["top"].min()),
            "w": int((sub["left"] + sub["width"]).max() - sub["left"].min()),
            "h": int(sub["height"].max()),
        })
    lines.sort(key=lambda d: (d["y"], d["x"]))
    return lines


_GENRE_FROM_RATING_RE = re.compile(
    r"\b\d[,.]\d\s*[«»·•+]+\s*([A-Z][^«»·•+]{2,40}?)\s*[«»·•+]"
)


def extract_genre_from_rating(text: str) -> str:
    """From a rating row like '4.7 · Mystery & Thriller · Jul 1, 2025 · 9 hr',
    return 'Mystery & Thriller'. Empty string if it doesn't match."""
    m = _GENRE_FROM_RATING_RE.search(text)
    if not m:
        return ""
    g = m.group(1).strip()
    # OCR sometimes glues a leading glyph to the genre: 'Happiness & Success'
    g = re.sub(r"^[^A-Za-z]+", "", g).strip()
    return g


_FAKE_ANCHOR_AUTHORS = {"charts", "top", "popular", "new"}
_SIDEBAR_TITLE_RE = re.compile(
    r"(?i:Playlist)\s*[»>·\-]+\s*\w+\s*[:\s\-]*([A-Z].+)$"
)
# Rating row, e.g. "4.8 « Happiness & Success « Oct 16, 2018 « 5 hr 36 min".
# We use this to detect the bottom of the previous book's row.
_RATING_ROW_RE = re.compile(
    r"\b\d[,.]\d\s*[«»·•+]|"          # "4,8 ·" / "4.5 +"
    r"\d+\s*hr.{0,5}min|"             # "5 hr 35 min", "1 hrf min"
    r"\b\d+\s*hr\s*\d+|"               # "5 hr35"
    r"\bShr\s*\d+|"                    # OCR sometimes reads "5 hr" as "Shr"
    r"\b\d+\s*min\b",                  # "35 min"
    re.IGNORECASE,
)


def parse_page(lines: list[dict], x_tol: int = 50) -> list[tuple[str, str]]:
    """For each 'Audiobook · Author' line, walk upward collecting same-x lines
    (the title rows), ignoring cover-art-column interlopers in between. Falls
    back to sidebar-fused-title extraction when nothing same-x is found."""
    pairs: list[tuple[str, str]] = []
    for i, ln in enumerate(lines):
        m = AUDIOBOOK_RE.match(ln["text"])
        if not m:
            continue
        author = clean_author(m.group(1))
        if not author or author.lower() in _FAKE_ANCHOR_AUTHORS:
            continue
        anchor_x = ln["x"]
        title_parts: list[str] = []
        # 1) same-x scan upward; stop when we hit the previous row's metadata
        for j in range(i - 1, -1, -1):
            prev = lines[j]
            if prev["y"] >= ln["y"]:
                continue
            if AUDIOBOOK_RE.match(prev["text"]):
                break
            if _RATING_ROW_RE.search(prev["text"]):
                break  # crossed into previous book's rating row
            if title_parts and ln["y"] - prev["y"] > 160:
                break
            if not title_parts and ln["y"] - prev["y"] > 180:
                break
            if abs(prev["x"] - anchor_x) > x_tol:
                continue  # cover-art-column line; skip but keep scanning
            title_parts.insert(0, prev["text"])
            if len(title_parts) >= 3:
                break
        title = clean_title(" ".join(title_parts))
        # 2) sidebar-fused fallback: when title got OCR'd onto a sidebar line
        if not title:
            for j in range(i - 1, -1, -1):
                prev = lines[j]
                if prev["y"] >= ln["y"]:
                    continue
                if ln["y"] - prev["y"] > 400:
                    break
                m2 = _SIDEBAR_TITLE_RE.search(prev["text"])
                if m2:
                    title = clean_title(m2.group(1))
                    if title:
                        break
        # 3) loose-x fallback: take the most-recent reasonably-long line within
        #    ~200px above the anchor whose x is within 200 of the anchor; skip
        #    lines that look like description body / review quotes
        if not title:
            for j in range(i - 1, -1, -1):
                prev = lines[j]
                if prev["y"] >= ln["y"]:
                    continue
                if ln["y"] - prev["y"] > 220:
                    break
                if abs(prev["x"] - anchor_x) > 200:
                    continue
                txt = prev["text"]
                if _RATING_ROW_RE.search(txt):
                    break
                # skip review-quote / description lines (smart quotes only)
                if "“" in txt or "”" in txt or "” —" in txt:
                    continue
                if txt.endswith(".") and len(txt.split()) > 12:
                    continue  # likely description prose
                cand = clean_title(txt)
                if cand and len(cand) > 8:
                    title = cand
                    break
        if not title:
            continue
        pairs.append((title, author))
    return pairs


def ocr_pdf_entries(pdf_path: Path, resolution: int = 200) -> tuple[list[tuple[str, str]], str]:
    """Returns (pairs, sheet_name). sheet_name is the most common genre across
    rating rows in the PDF, falling back to '' if no genres detected."""
    out: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()
    genre_counts: dict[str, int] = {}
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            im = page.to_image(resolution=resolution).original
            lines = ocr_page_lines(im)
            for ln in lines:
                g = extract_genre_from_rating(ln["text"])
                if g:
                    genre_counts[g] = genre_counts.get(g, 0) + 1
            for t, a in parse_page(lines):
                key = (t.lower(), a.lower())
                if key in seen:
                    continue
                seen.add(key)
                out.append((t, a))
    sheet_name = max(genre_counts, key=genre_counts.get) if genre_counts else ""
    return out, sheet_name


_ICON_TOKEN = re.compile(
    r"^("
    r"[®©@*]+"                 # single symbols
    r"|fa[\)\]]?"              # 'fa]' font-awesome alt-text leak
    r"|[a-zA-Z]{1,3}[\)\]]"    # 're)', 'a]', 'Ca]', 'iS)', 'Ca)'
    r"|\([a-zA-Z]{1,3}\)"      # '(c)', '(Ca)'
    r"|[0oO]"                  # stray 'O'/'0' from circular icons
    r"|[\W_]+"                 # punctuation-only token
    r")$",
    re.IGNORECASE,
)


def clean_title(t: str) -> str:
    # strip leading & trailing row-icon OCR artifacts: "® fa]", "@ 0", "ic) fa]", etc.
    tokens = t.split()
    while tokens:
        if _ICON_TOKEN.match(tokens[0]):
            tokens.pop(0)
            continue
        # also strip a short alphabetic prefix when the NEXT token is an icon
        # (handles cover-art bleed like 'eon Ca)' or 'TR INA. Dear Debbie')
        if (len(tokens) >= 2 and len(tokens[0]) <= 4
                and _ICON_TOKEN.match(tokens[1])):
            tokens.pop(0)
            continue
        break
    while tokens and _ICON_TOKEN.match(tokens[-1]):
        tokens.pop()
    t = " ".join(tokens)
    # strip leading list numbering
    t = re.sub(r"^\s*\d+[\.\)]\s*", "", t)
    # strip cover-art glyph prefixes like '38")', '(c)', '7"'
    t = re.sub(r'^\d{1,3}["\')\]\}]+\s+', "", t)
    # strip leading single-digit cover-art noise: '6 Atomic Habits' -> 'Atomic Habits'
    t = re.sub(r"^\d{1,2}\s+(?=[A-Z])", "", t)
    # collapse cover-art doubling: 'THINK Think and Grow Rich' -> 'Think and Grow Rich'
    t = re.sub(r"^(\w+)\s+(\1)\b", r"\2", t, flags=re.IGNORECASE)
    # strip Spotify banner caption fusion at line start
    t = re.sub(
        r"^(Spo[tk]i?fy?\b.*?\b\d+\s+books?\s*"
        r"|Top\s+[\w&,\s]+?audiobooks?[\w,\s]*?\b"
        r"|updated\s+\w+\s+[\w«»·•+]*)\s*",
        "", t, flags=re.IGNORECASE,
    )
    # OCR sometimes glues a leading single letter to the next word:
    # 'ACourt of Wings' → 'A Court of Wings'.
    t = re.sub(r"^([A-Z])([A-Z][a-z])", r"\1 \2", t)
    t = t.strip(" \t·•+-—_")
    if len(t) < 3:
        return ""
    return t


def clean_author(a: str) -> str:
    a = a.strip(" \t·•+-—_,")
    a = re.sub(r"\s{2,}", " ", a)
    a = re.split(r"[;|]", a)[0].strip()
    return a


def _get_with_retries(session, url, params, max_retries=4):
    """GET with exponential backoff for 429/5xx."""
    last = None
    for attempt in range(max_retries):
        try:
            r = session.get(url, params=params, timeout=20)
            if r.status_code in (429, 500, 502, 503, 504):
                last = f"HTTP {r.status_code}"
                wait = 2 ** attempt + 0.5
                time.sleep(wait)
                continue
            r.raise_for_status()
            return r.json()
        except requests.HTTPError as e:
            last = str(e)
            if r.status_code < 500 and r.status_code != 429:
                break
            time.sleep(2 ** attempt + 0.5)
        except Exception as e:
            last = str(e)
            time.sleep(2 ** attempt + 0.5)
    raise RuntimeError(f"giving up after {max_retries}: {last}")


def google_books_lookup(title: str, author: str, session: requests.Session,
                        api_key: str = "") -> dict:
    """Query Google Books for publisher/publishedDate. Returns empty dict on failure."""
    queries = [
        {"q": f'intitle:"{title}" inauthor:"{author}"'},
        {"q": f"{title} {author}"},
    ]
    items = []
    for q in queries:
        params = {**q, "printType": "books", "maxResults": 40, "orderBy": "relevance"}
        if api_key:
            params["key"] = api_key
        try:
            data = _get_with_retries(session, GOOGLE_BOOKS, params)
        except Exception as e:
            return {"_error": f"google_books: {e}"}
        items = data.get("items") or []
        if items:
            break
    if not items:
        return {}

    a_low = author.lower()
    matches = [it for it in items
               if _author_matches(a_low,
                                  it.get("volumeInfo", {}).get("authors") or [])]
    if not matches:
        # No author match — refuse to invent a result. Let the fallback try.
        return {}

    def pubdate(it):
        return (it.get("volumeInfo", {}).get("publishedDate") or "9999")

    earliest = min(matches, key=pubdate)
    publisher = ""
    for it in matches:
        p = it.get("volumeInfo", {}).get("publisher")
        if p:
            publisher = p
            break
    vi = earliest.get("volumeInfo", {})
    return {
        "publisher": publisher,
        "first_published": vi.get("publishedDate", ""),
        "matched_title": vi.get("title", ""),
        "matched_authors": ", ".join(vi.get("authors") or []),
        "source": "google_books",
    }


def _author_matches(queried_low: str, candidate_authors: list[str]) -> bool:
    """True if any candidate author shares the last-name token with our query.
    Handles 'J.RR. Tolkien' vs 'John Ronald Reuel Tolkien' (last-token match)."""
    if not candidate_authors:
        return False
    q_tokens = [t for t in re.split(r"[^a-zA-Z]+", queried_low) if len(t) > 2]
    if not q_tokens:
        return any(queried_low in (c or "").lower() for c in candidate_authors)
    q_last = q_tokens[-1]
    for c in candidate_authors:
        c_low = (c or "").lower()
        if queried_low in c_low or c_low in queried_low:
            return True
        if q_last and q_last in c_low:
            return True
    return False


def _short_title(t: str) -> str:
    """Trim a long subtitle: 'Atomic Habits: An Easy...' → 'Atomic Habits'."""
    return re.split(r"[:\-—–]", t, maxsplit=1)[0].strip()


def open_library_lookup(title: str, author: str, session: requests.Session) -> dict:
    """Open Library. Tries several query variants for coverage."""
    short = _short_title(title)
    attempts = [
        {"title": title, "author": author, "limit": 10},
        {"title": short, "author": author, "limit": 10},
        {"q": f"{short} {author}", "limit": 10},
    ]
    docs = []
    for params in attempts:
        try:
            data = _get_with_retries(session, OPEN_LIBRARY, params)
        except Exception as e:
            return {"_error": f"open_library: {e}"}
        docs = data.get("docs") or []
        if docs:
            break
    if not docs:
        return {}

    a_low = author.lower()
    # Prefer docs whose authors match
    ranked = sorted(
        docs,
        key=lambda d: (
            0 if any(a_low in (n or "").lower() for n in (d.get("author_name") or [])) else 1,
            d.get("first_publish_year") or 9999,
        ),
    )
    d = ranked[0]
    publishers = d.get("publisher") or []
    return {
        "publisher": publishers[0] if publishers else "",
        "first_published": str(d.get("first_publish_year") or ""),
        "matched_title": d.get("title", ""),
        "matched_authors": ", ".join(d.get("author_name") or []),
        "source": "open_library",
    }


def enrich(title: str, author: str, session: requests.Session,
           api_key: str, prefer: str) -> dict:
    """`prefer` modes:
       - 'google'     : Google primary, OL fallback
       - 'openlibrary': OL primary, Google fallback
       - 'hybrid'     : Google for publisher, OL for first_publish_year (true original);
                        falls back to the other when one side is empty.
    """
    if prefer == "hybrid":
        g = google_books_lookup(title, author, session, api_key=api_key)
        if g.get("_error"):
            g = {}
        o = open_library_lookup(title, author, session)
        if o.get("_error"):
            o = {}
        publisher = g.get("publisher") or o.get("publisher", "")
        first = o.get("first_published") or g.get("first_published", "")
        # source label tells caller which side won the date battle
        date_src = "open_library" if o.get("first_published") else ("google_books" if g.get("first_published") else "")
        pub_src = "google_books" if g.get("publisher") else ("open_library" if o.get("publisher") else "")
        if not publisher and not first:
            return {"publisher": "", "first_published": "", "matched_title": "",
                    "matched_authors": "", "source": ""}
        return {
            "publisher": publisher,
            "first_published": first,
            "matched_title": g.get("matched_title") or o.get("matched_title", ""),
            "matched_authors": g.get("matched_authors") or o.get("matched_authors", ""),
            "source": f"pub:{pub_src};date:{date_src}",
        }

    order = ["google_books", "open_library"] if prefer == "google" else ["open_library", "google_books"]
    best = {}
    for src in order:
        if src == "google_books":
            res = google_books_lookup(title, author, session, api_key=api_key)
        else:
            res = open_library_lookup(title, author, session)
        if res.get("_error"):
            continue
        if res.get("publisher") and res.get("first_published"):
            return res
        if (res.get("publisher") or res.get("first_published")) and not best:
            best = res
    return best or {"publisher": "", "first_published": "", "matched_title": "",
                    "matched_authors": "", "source": ""}


_PLAYLIST_ID_RE = re.compile(r"playlist-([A-Za-z0-9]{15,})")


def derive_sheet_name(pdf_name: str, banner: str = "") -> str:
    """Banner-based naming with playlist-ID fallback."""
    if banner:
        return banner
    m = _PLAYLIST_ID_RE.search(pdf_name)
    return m.group(1) if m else Path(pdf_name).stem


def safe_sheet_name(name: str, used: set[str]) -> str:
    bad = set(r'[]:*?/\\')
    cleaned = "".join("_" if c in bad else c for c in name)[:31] or "Sheet"
    base, n = cleaned, 1
    while cleaned in used:
        suffix = f"_{n}"
        cleaned = (base[: 31 - len(suffix)]) + suffix
        n += 1
    used.add(cleaned)
    return cleaned


def write_workbook(out_path: Path, per_pdf: dict[str, dict],
                   append_to: Path | None = None,
                   preserve: set[str] | None = None) -> None:
    """If append_to is provided, load that workbook, drop sheets not in `preserve`,
    and add the new sheets. Otherwise build a fresh workbook."""
    preserve = preserve or set()
    if append_to and append_to.exists():
        from openpyxl import load_workbook
        wb = load_workbook(append_to)
        for sn in list(wb.sheetnames):
            if sn not in preserve:
                del wb[sn]
        used = set(wb.sheetnames)
    else:
        wb = Workbook()
        wb.remove(wb.active)
        used = set()

    headers = ["Title", "Author", "Publisher (print)", "First Published (print)",
               "Matched Title", "Matched Authors", "Source"]
    for pdf_name, payload in per_pdf.items():
        rows = payload["rows"]
        banner = payload.get("banner", "")
        ws = wb.create_sheet(safe_sheet_name(derive_sheet_name(pdf_name, banner), used))
        ws.append(headers)
        for row in rows:
            ws.append([
                row["title"], row["author"],
                row.get("publisher", ""), row.get("first_published", ""),
                row.get("matched_title", ""), row.get("matched_authors", ""),
                row.get("source", ""),
            ])
        # autosize-ish
        for col_idx, h in enumerate(headers, start=1):
            max_len = max([len(str(h))] + [len(str(ws.cell(r, col_idx).value or ""))
                                           for r in range(2, ws.max_row + 1)])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
        ws.freeze_panes = "A2"
    wb.save(out_path)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("out_xlsx", type=Path, help="Output .xlsx path")
    ap.add_argument("pdfs", type=Path, nargs="+", help="One or more PDF screencaptures")
    ap.add_argument("--resolution", type=int, default=200,
                    help="OCR rendering DPI (default 200)")
    ap.add_argument("--sleep", type=float, default=0.4,
                    help="Sleep between API calls (default 0.4s)")
    ap.add_argument("--prefer", choices=["google", "openlibrary", "hybrid"], default="google",
                    help="google = Google primary w/ OL fallback; openlibrary = reverse; "
                         "hybrid = Google publisher + Open Library first-publish year")
    ap.add_argument("--google-api-key", default="",
                    help="Google Books API key (else uses google_books_key.txt next to this script)")
    ap.add_argument("--append-to", type=Path, default=None,
                    help="Existing xlsx to merge into; replaces sheets not in --preserve")
    ap.add_argument("--preserve", action="append", default=[],
                    help="Sheet name to keep when --append-to is set (can repeat)")
    args = ap.parse_args(argv)

    if not args.google_api_key:
        key_file = Path(__file__).parent / "google_books_key.txt"
        if key_file.exists():
            args.google_api_key = key_file.read_text().strip()

    session = requests.Session()
    per_pdf: dict[str, dict] = {}

    for pdf in args.pdfs:
        if not pdf.exists():
            print(f"[skip] {pdf} does not exist", file=sys.stderr)
            continue
        print(f"[ocr]  {pdf.name}", file=sys.stderr)
        pairs, banner = ocr_pdf_entries(pdf, resolution=args.resolution)
        print(f"       banner: {banner!r}  parsed {len(pairs)} entries",
              file=sys.stderr)
        rows: list[dict] = []
        for title, author in pairs:
            print(f"[meta] {title!r:60.60} | {author!r}", file=sys.stderr)
            meta = enrich(title, author, session,
                          api_key=args.google_api_key, prefer=args.prefer)
            rows.append({"title": title, "author": author, **meta})
            time.sleep(args.sleep)
        per_pdf[pdf.name] = {"rows": rows, "banner": banner}

    args.out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    write_workbook(args.out_xlsx, per_pdf,
                   append_to=args.append_to,
                   preserve=set(args.preserve))
    print(f"[done] wrote {args.out_xlsx}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
