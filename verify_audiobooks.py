#!/usr/bin/env python3
"""
Per-row verification pass over audiobooks.xlsx.

For every row in non-preserved sheets:
  1. Fix obvious OCR typos in the author cell.
  2. Re-query Google Books with the cleaned (title, author).
  3. If our current date is year-only and Google has a stricter date for the
     same book in the same year (or earlier with a known publisher), upgrade it.
  4. Flag rows where Google's matched authors don't include our author —
     these get a 'review' note in Source.

Writes back in place. Prints a change log.
"""
import re
import sys
import time
from pathlib import Path

import requests
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from spotify_scrape import (
    GOOGLE_BOOKS, _get_with_retries, _author_matches, _short_title,
)

PRESERVE = {"Top", "Romance"}


def fix_ocr_author(a: str) -> str:
    """Repair common OCR mistakes in author cells."""
    if not a:
        return a
    s = a
    # OCR commonly reads '.' as ',' inside initials: "C. S, Lewis" -> "C. S. Lewis"
    s = re.sub(r"\b([A-Z])\s*,\s*([A-Z][a-z])", r"\1. \2", s)
    # 'J.RR. Tolkien' -> 'J.R.R. Tolkien'
    s = re.sub(r"\b([A-Z])\.([A-Z]{2,})\b", lambda m: m.group(1) + "." + ".".join(m.group(2)) + ".", s)
    # Collapse double spaces
    s = re.sub(r"\s{2,}", " ", s).strip()
    # Strip stray trailing punctuation
    s = s.rstrip(",;:")
    return s


def is_year_only(d: str) -> bool:
    return bool(re.fullmatch(r"\d{4}", (d or "").strip()))


def fmt_date(d: str) -> str:
    return (d or "").strip()


def google_lookup_strict(title: str, author: str, session, api_key: str) -> list[dict]:
    queries = [
        {"q": f'intitle:"{title}" inauthor:"{author}"'},
        {"q": f'intitle:"{_short_title(title)}" inauthor:"{author}"'},
        {"q": f'"{_short_title(title)}" "{author}"'},
    ]
    for q in queries:
        params = {**q, "printType": "books", "maxResults": 40, "key": api_key}
        try:
            data = _get_with_retries(session, GOOGLE_BOOKS, params)
        except Exception:
            continue
        items = data.get("items") or []
        if items:
            return items
    return []


def best_match(items, author):
    a_low = author.lower()
    matches = [
        it for it in items
        if _author_matches(a_low, it.get("volumeInfo", {}).get("authors") or [])
    ]
    return matches


def refine_row(title, author, publisher, date, matched_title, matched_authors,
               source, session, api_key):
    """Return (new_publisher, new_date, new_matched_title, new_matched_authors,
                new_source, change_note). 'change_note' is '' if nothing changed."""
    clean = fix_ocr_author(author)
    new_author = clean if clean != author else author

    items = google_lookup_strict(title, clean, session, api_key)
    if not items:
        note = "no-results-on-recheck" if not publisher and not date else ""
        return (publisher, date, matched_title, matched_authors,
                source + (";flag:no-google-hit" if note and "flag:" not in source else ""),
                new_author, note)

    matches = best_match(items, clean)
    if not matches:
        flag = "author-mismatch"
        # Look at top candidate's authors for context
        top_authors = ", ".join(items[0].get("volumeInfo", {}).get("authors") or [])
        new_source = source if "flag:" in source else source + f";flag:{flag}({top_authors[:40]})"
        return (publisher, date, matched_title, matched_authors,
                new_source, new_author, flag)

    note_parts = []

    # If date is year-only, try to find a fuller date for the same edition
    if is_year_only(date):
        same_year = [
            it for it in matches
            if (it.get("volumeInfo", {}).get("publishedDate") or "").startswith(date)
            and len(it.get("volumeInfo", {}).get("publishedDate", "")) >= 10
        ]
        # Prefer one whose publisher matches our current publisher (likely same edition)
        if publisher and same_year:
            same_pub = [
                it for it in same_year
                if publisher.lower() in (it.get("volumeInfo", {}).get("publisher") or "").lower()
                or (it.get("volumeInfo", {}).get("publisher") or "").lower() in publisher.lower()
            ]
            picked = (same_pub or same_year)[0]
        elif same_year:
            picked = same_year[0]
        else:
            picked = None
        if picked:
            new_date = picked["volumeInfo"]["publishedDate"]
            new_pub = picked["volumeInfo"].get("publisher", publisher) or publisher
            note_parts.append(f"date {date} -> {new_date}")
            if new_pub != publisher:
                note_parts.append(f"pub {publisher!r} -> {new_pub!r}")
            return (new_pub, new_date,
                    picked["volumeInfo"].get("title", matched_title),
                    ", ".join(picked["volumeInfo"].get("authors") or []),
                    source, new_author, "; ".join(note_parts))

    # If publisher empty and we have a confident match with a publisher, fill it
    if not publisher:
        for it in matches:
            p = it.get("volumeInfo", {}).get("publisher")
            if p:
                note_parts.append(f"publisher filled: {p!r}")
                return (p, date, matched_title, matched_authors,
                        source, new_author, "; ".join(note_parts))

    if new_author != author:
        note_parts.append(f"author cleaned {author!r} -> {new_author!r}")
    return (publisher, date, matched_title, matched_authors,
            source, new_author, "; ".join(note_parts))


def main(xlsx_path: str):
    p = Path(xlsx_path)
    wb = load_workbook(p)

    api_key_file = Path(__file__).parent / "google_books_key.txt"
    api_key = api_key_file.read_text().strip() if api_key_file.exists() else ""
    session = requests.Session()

    changes = 0
    flagged = 0
    for sn in wb.sheetnames:
        if sn in PRESERVE:
            continue
        ws = wb[sn]
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        col = {h: i for i, h in enumerate(header)}
        i_title = col["Title"]
        i_auth = col["Author"]
        i_pub = col["Publisher (print)"]
        i_date = col["First Published (print)"]
        i_mt = col.get("Matched Title", -1)
        i_ma = col.get("Matched Authors", -1)
        i_src = col.get("Source", -1)

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            title = str(row[i_title].value or "")
            author = str(row[i_auth].value or "")
            publisher = str(row[i_pub].value or "")
            date = fmt_date(str(row[i_date].value or ""))
            mt = str(row[i_mt].value or "") if i_mt >= 0 else ""
            ma = str(row[i_ma].value or "") if i_ma >= 0 else ""
            src = str(row[i_src].value or "") if i_src >= 0 else ""

            new_pub, new_date, new_mt, new_ma, new_src, new_author, note = refine_row(
                title, author, publisher, date, mt, ma, src, session, api_key
            )
            time.sleep(0.15)
            if not note and new_author == author:
                continue

            if new_author != author:
                row[i_auth].value = new_author
            if new_pub != publisher:
                row[i_pub].value = new_pub
            if new_date != date:
                row[i_date].value = new_date
            if i_mt >= 0 and new_mt and new_mt != mt:
                row[i_mt].value = new_mt
            if i_ma >= 0 and new_ma and new_ma != ma:
                row[i_ma].value = new_ma
            if i_src >= 0 and new_src != src:
                row[i_src].value = new_src

            if "flag:" in (new_src or ""):
                flagged += 1
            changes += 1
            print(f"[{sn} r{row_idx}] {title[:35]!r} | {note}", file=sys.stderr)

    wb.save(p)
    print(f"\nDone. rows changed={changes}  flagged={flagged}", file=sys.stderr)


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "audiobooks.xlsx")
