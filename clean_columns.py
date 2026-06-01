#!/usr/bin/env python3
"""
Drop Matched Title / Matched Authors / Source columns from every data sheet.
Replace the existing N (total count) column with a serial-number N as the
FIRST column of every sheet (1..n).
"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

XLSX = Path(__file__).parent / "audiobooks.xlsx"
DROP_COLS = {"Matched Title", "Matched Authors", "Source"}


def clean_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    header = list(rows[0])

    # Keep indices for columns we are NOT dropping AND that aren't the old N
    keep_idx = [
        i for i, h in enumerate(header)
        if h is not None and h not in DROP_COLS and h != "N"
    ]
    new_header = ["N"] + [header[i] for i in keep_idx]
    new_rows = [new_header]
    serial = 0
    for r in rows[1:]:
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        serial += 1
        new_rows.append([serial] + [r[i] if i < len(r) else None for i in keep_idx])

    ws.delete_rows(1, ws.max_row)
    for row in new_rows:
        ws.append(row)

    # Widths
    widths = [6, 42, 24, 26, 22, 16, 10]
    for cidx, w in enumerate(widths[: len(new_header)], start=1):
        ws.column_dimensions[get_column_letter(cidx)].width = w
    ws.freeze_panes = "A2"


def main():
    wb = load_workbook(XLSX)
    for sn in wb.sheetnames:
        if sn == "Summary":
            continue
        clean_sheet(wb[sn])
        print(f"  cleaned {sn}")
    wb.save(XLSX)
    print(f"\nSaved {XLSX.name}")


if __name__ == "__main__":
    main()
