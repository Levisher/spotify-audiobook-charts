#!/usr/bin/env python3
"""
Replace the 'First Published (print)' column in an existing audiobooks.xlsx
with Open Library's first_publish_year (the true original publication year),
falling back to the existing Google Books date when OL has nothing.

Usage:
    python3 patch_dates_from_openlibrary.py audiobooks.xlsx
"""
import sys
import time
from pathlib import Path

import requests
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from spotify_scrape import open_library_lookup  # noqa: E402


def _year_of(s: str) -> str:
    s = (s or "").strip()
    return s[:4] if len(s) >= 4 and s[:4].isdigit() else ""


def main(xlsx_path: str, sleep: float = 0.3):
    p = Path(xlsx_path)
    wb = load_workbook(p)
    session = requests.Session()

    total_rows = 0
    replaced = 0
    unchanged_year = 0
    no_ol_data = 0

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=False))
        if not rows:
            continue
        header = [c.value for c in rows[0]]
        try:
            i_title = header.index("Title")
            i_author = header.index("Author")
            i_date = header.index("First Published (print)")
            i_source = header.index("Source") if "Source" in header else None
        except ValueError as e:
            print(f"[skip] {sn}: missing column {e}", file=sys.stderr)
            continue

        for row in rows[1:]:
            total_rows += 1
            title = row[i_title].value or ""
            author = row[i_author].value or ""
            current = str(row[i_date].value or "")
            ol = open_library_lookup(title, author, session)
            time.sleep(sleep)
            ol_year = ol.get("first_published", "")
            if not ol_year:
                no_ol_data += 1
                continue
            if _year_of(current) == ol_year:
                unchanged_year += 1
                continue
            print(f"[{sn}] {title[:45]!r:48} | {current!r:14} -> {ol_year!r}", file=sys.stderr)
            row[i_date].value = ol_year
            if i_source is not None:
                cur_src = str(row[i_source].value or "")
                row[i_source].value = (cur_src + ";date:open_library").lstrip(";")
            replaced += 1

    wb.save(p)
    print(f"\nDone. rows={total_rows}  replaced={replaced}  "
          f"same-year={unchanged_year}  no-OL-data={no_ol_data}",
          file=sys.stderr)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__, file=sys.stderr)
        raise SystemExit(2)
    main(sys.argv[1])
